import streamlit as st
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
import pandas as pd
import time
import io

# ================= 사용자 설정 =================
URL = "https://coupon.withhive.com/soulstrike?t=1737278324092"
CS_CODE_INPUT_SELECTOR = "input#cs_code"
COUPON_INPUT_SELECTOR = "input#coupon_code"
SUBMIT_BUTTON_SELECTOR = "button.btn_use"
POPUP_SELECTOR = "div.pop_wrap.coupon_lyr"
POPUP_MSG_ID = "layer_msg"
POPUP_CLOSE_ID = "layer_close_btn"
SLEEP_SEC = 1.5
# =================================================

st.title("SoulStrike 쿠폰 자동등록기 (웹버전)")

# 엑셀 업로드
uploaded_file = st.file_uploader("엑셀 파일 선택 (.xlsx)", type="xlsx")
coupon_code = st.text_input("쿠폰 코드 입력")

# 결과 출력용
log_placeholder = st.empty()
logs = []

def log(msg):
    logs.append(msg)
    log_placeholder.text_area("실행 로그", value="\n".join(logs), height=400)

def run_coupon_process(excel_file, coupon_code):
    try:
        wb = load_workbook(excel_file)
        ws = wb.active
        ids = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value]

        options = Options()
        options.add_argument("--headless")  # 브라우저 창 안 띄우기
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")

        log("✅ Chrome 실행 중...")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
        driver.get(URL)
        wait = WebDriverWait(driver, 10)
        log("✅ 페이지 접속 완료.")

        results = []

        for idx, cs_code in enumerate(ids, start=2):
            log(f"[{idx-1}] {cs_code} 처리 중...")
            try:
                # CS 코드 입력
                cs_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, CS_CODE_INPUT_SELECTOR)))
                cs_input.clear()
                cs_input.send_keys(str(cs_code))

                # 쿠폰 코드 입력
                coupon_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, COUPON_INPUT_SELECTOR)))
                coupon_input.clear()
                coupon_input.send_keys(coupon_code)

                # 제출 버튼 클릭
                submit_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, SUBMIT_BUTTON_SELECTOR)))
                submit_btn.click()

                # 팝업 확인
                try:
                    popup = WebDriverWait(driver, 5).until(
                        EC.visibility_of_element_located((By.CSS_SELECTOR, POPUP_SELECTOR))
                    )
                    msg_elem = popup.find_element(By.ID, POPUP_MSG_ID)
                    msg = msg_elem.text

                    if "완료" in msg or "쿠폰함" in msg:
                        results.append(f"✅ {cs_code} → 성공: {msg}")
                    else:
                        results.append(f"⚠️ {cs_code} → 실패: {msg}")

                    # 팝업 닫기
                    close_btn = popup.find_element(By.ID, POPUP_CLOSE_ID)
                    close_btn.click()

                except:
                    results.append(f"⚠️ {cs_code} → 팝업 없음/확인 필요")

                # 로그 갱신
                log("\n".join(results))
                time.sleep(SLEEP_SEC)

            except Exception as e:
                results.append(f"❌ {cs_code} 처리 중 오류: {e}")
                log("\n".join(results))

        driver.quit()
        log("\n✅ 모든 처리 완료!")

        # 최종 결과를 파일로도 제공 (선택)
        df = pd.DataFrame({"CS코드": ids, "결과": results})
        output = io.BytesIO()
        df.to_excel(output, index=False)
        st.download_button(
            label="결과 다운로드 (.xlsx)",
            data=output.getvalue(),
            file_name="coupon_result.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        )

    except Exception as e:
        log(f"❌ 오류 발생: {e}")

if st.button("실행"):
    if uploaded_file is None:
        st.warning("엑셀 파일을 선택하세요.")
    elif not coupon_code.strip():
        st.warning("쿠폰 코드를 입력하세요.")
    else:
        run_coupon_process(uploaded_file, coupon_code.strip())
