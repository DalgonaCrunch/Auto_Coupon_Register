import tkinter as tk
from tkinter import filedialog, messagebox
from selenium import webdriver
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.chrome.options import Options
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import load_workbook
from screeninfo import get_monitors
import threading
import time

URL = "https://coupon.withhive.com/soulstrike?t=1737278324092"

CS_CODE_INPUT_SELECTOR = "input#cs_code"
COUPON_INPUT_SELECTOR = "input#coupon_code"
SUBMIT_BUTTON_SELECTOR = "button.btn_use"
SLEEP_SEC = 1.5

def log_append(msg):
    txt_log.insert(tk.END, msg + "\n")
    txt_log.see(tk.END)
    root.update_idletasks()

def run_coupon_process(excel_path, coupon_codes):
    try:
        wb = load_workbook(excel_path)
        ws = wb.active
        ids = [row[0].value for row in ws.iter_rows(min_row=2) if row[0].value]

        options = Options()
        options.add_argument("--no-sandbox")
        options.add_argument("--disable-dev-shm-usage")
        options.add_argument("--disable-gpu")

        log_append("✅ Chrome 실행 중...")
        driver = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)

        monitor = get_monitors()[0]
        screen_width = monitor.width
        screen_height = monitor.height

        win_width = 100
        win_height = 100
        pos_x = screen_width - win_width - 20
        pos_y = screen_height - win_height - 80

        driver.set_window_size(win_width, win_height)
        driver.set_window_position(pos_x, pos_y)
        driver.get(URL)
        wait = WebDriverWait(driver, 10)
        log_append("✅ 페이지 접속 완료.")

        for row_idx, cs_code in enumerate(ids, start=2):
            lbl_status.config(text=f"현재 처리 중: {cs_code}")

            for coupon_index, coupon in enumerate(coupon_codes):
                try:
                    log_append(f"▶ {cs_code} / 쿠폰: {coupon}")

                    cs_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, CS_CODE_INPUT_SELECTOR)))
                    cs_input.clear()
                    cs_input.send_keys(str(cs_code))

                    coupon_input = wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, COUPON_INPUT_SELECTOR)))
                    coupon_input.clear()
                    coupon_input.send_keys(coupon)

                    submit_btn = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, SUBMIT_BUTTON_SELECTOR)))
                    submit_btn.click()

                    # ✅ 결과 기록할 열 : C=3, D=4, E=5 ...
                    col = 3 + coupon_index
                    
                    try:
                        popup = WebDriverWait(driver, 5).until(
                            EC.visibility_of_element_located((By.CSS_SELECTOR, "div.pop_wrap.coupon_lyr"))
                        )
                        msg_elem = popup.find_element(By.ID, "layer_msg")
                        msg = msg_elem.text

                        ws.cell(row=row_idx, column=col, value=msg)
                        log_append(f"✅ 결과: {msg}")

                        close_btn = popup.find_element(By.ID, "layer_close_btn")
                        close_btn.click()

                    except:
                        ws[f"C{row_idx}"] = "오류: 팝업 없음"
                        log_append("⚠ 팝업 없음")

                    time.sleep(SLEEP_SEC)

                except Exception as e:
                    log_append(f"❌ 오류: {e}")

        result_path = excel_path.replace(".xlsx", "_result.xlsx")
        wb.save(result_path)
        driver.quit()

        lbl_status.config(text="✅ 완료")
        messagebox.showinfo("완료", f"모든 코드 처리됨!\n결과: {result_path}")

    except Exception as e:
        messagebox.showerror("오류 발생", str(e))

def start_process():
    excel_path = entry_excel.get()
    coupon_codes = [e.get().strip() for e in entry_coupon_list if e.get().strip()]

    if not excel_path:
        messagebox.showwarning("경고", "엑셀 파일을 선택하세요.")
        return
    if not coupon_codes:
        messagebox.showwarning("경고", "최소 1개의 쿠폰 코드를 입력하세요.")
        return

    threading.Thread(target=run_coupon_process, args=(excel_path, coupon_codes), daemon=True).start()

def browse_excel():
    path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
    if path:
        entry_excel.delete(0, tk.END)
        entry_excel.insert(0, path)

def add_coupon_field():
    entry = tk.Entry(frame, width=35)
    entry_coupon_list.append(entry)
    entry.grid(row=3 + len(entry_coupon_list), column=0, pady=3, sticky="w")

root = tk.Tk()
root.title("SoulStrike 쿠폰 자동등록기 - 다중 쿠폰")
root.geometry("600x500")
root.resizable(False, False)

frame = tk.Frame(root, padx=20, pady=20)
frame.pack(fill="both", expand=True)

tk.Label(frame, text="엑셀 파일 경로:").grid(row=0, column=0, sticky="w")
entry_excel = tk.Entry(frame, width=45)
entry_excel.grid(row=1, column=0, pady=5, sticky="w")
tk.Button(frame, text="찾아보기", command=browse_excel).grid(row=1, column=1, padx=5)

tk.Label(frame, text="쿠폰 코드 목록 (여러개 입력 가능):").grid(row=2, column=0, pady=(15, 0), sticky="w")

entry_coupon_list = []
add_coupon_field()

tk.Button(frame, text="+ 쿠폰 추가", command=add_coupon_field).grid(row=3, column=1, padx=5)

tk.Button(frame, text="실행하기", command=start_process, bg="#3a7afe", fg="white", width=15).grid(row=20, column=0, pady=(15, 0))

lbl_status = tk.Label(frame, text="대기 중", fg="gray")
lbl_status.grid(row=21, column=0, pady=(10, 5), sticky="w")

tk.Label(frame, text="실행 로그:").grid(row=22, column=0, sticky="w")
txt_log = tk.Text(frame, height=10, width=70, wrap="word", bg="#f8f8f8")
txt_log.grid(row=23, column=0, columnspan=2, pady=5)
scrollbar = tk.Scrollbar(frame, command=txt_log.yview)
scrollbar.grid(row=23, column=2, sticky="ns")
txt_log.config(yscrollcommand=scrollbar.set)

root.mainloop()
