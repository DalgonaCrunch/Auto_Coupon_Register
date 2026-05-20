"""Telegram bot entry point: receive commands, run the coupon engine, reply."""
from __future__ import annotations

import asyncio
import json
import logging
import os
import re
from collections import Counter
from functools import wraps

from dotenv import load_dotenv
from telegram import Update
from telegram.constants import ParseMode
from telegram.ext import (
    Application,
    CommandHandler,
    ContextTypes,
    MessageHandler,
    filters,
)

import coupon_engine
import id_store
from paths import base_dir

# Load .env from the directory next to the exe / script, not CWD.
load_dotenv(dotenv_path=base_dir() / ".env")

logging.basicConfig(
    format="%(asctime)s %(levelname)s %(name)s - %(message)s",
    level=logging.INFO,
)
logger = logging.getLogger("coupon_bot")

TOKEN = os.environ.get("TELEGRAM_BOT_TOKEN", "").strip()
ALLOWED_CHAT_IDS: set[int] = {
    int(x) for x in os.environ.get("ALLOWED_CHAT_IDS", "").split(",") if x.strip()
}
DEFAULT_SERVER = os.environ.get("DEFAULT_SERVER", "KR/JP/GLB").strip()
HEADLESS = os.environ.get("HEADLESS", "1").strip() == "1"

STATE_PATH = base_dir() / "bot_state.json"
_run_lock = asyncio.Lock()


def _load_state() -> dict:
    if not STATE_PATH.exists():
        return {}
    try:
        return json.loads(STATE_PATH.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def _save_state(state: dict) -> None:
    STATE_PATH.write_text(json.dumps(state, ensure_ascii=False, indent=2), encoding="utf-8")


def current_server() -> str:
    return _load_state().get("server", DEFAULT_SERVER)


def _parse_args(message_text: str | None) -> list[str]:
    """Split the args of a /command message by whitespace, commas, and newlines.

    Drops the leading /command (or /command@botname) token. Dedupes while
    preserving order. Lets users paste many IDs/codes at once in any format:

        /add user1 user2,user3
        /add
        user1
        user2, user3
    """
    if not message_text:
        return []
    parts = message_text.split(None, 1)
    if len(parts) < 2:
        return []
    tokens = re.split(r"[,\s]+", parts[1])
    seen: set[str] = set()
    out: list[str] = []
    for t in tokens:
        t = t.strip()
        if t and t not in seen:
            out.append(t)
            seen.add(t)
    return out


def authorized(handler):
    @wraps(handler)
    async def wrapper(update: Update, context: ContextTypes.DEFAULT_TYPE):
        chat_id = update.effective_chat.id if update.effective_chat else None
        if chat_id not in ALLOWED_CHAT_IDS:
            logger.warning("Unauthorized chat_id=%s blocked", chat_id)
            if update.message:
                await update.message.reply_text(
                    f"권한이 없습니다. chat_id={chat_id} 를 관리자에게 전달하세요."
                )
            return
        return await handler(update, context)

    return wrapper


@authorized
async def cmd_start(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    await update.message.reply_text(
        "SoulStrike 쿠폰 봇입니다.\n/help 로 사용법을 확인하세요."
    )


@authorized
async def cmd_help(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    text = (
        "*사용법*\n"
        "`/coupon <코드> [코드 ...]` — 등록된 모든 ID에 쿠폰 적용\n"
        "`/add <ID> [ID ...]` — ID 추가\n"
        "`/del <ID> [ID ...]` — ID 삭제\n"
        "`/list` — 등록된 ID 목록\n"
        "`/server <KR/JP/GLB>` — 기본 서버 변경 (현재: "
        f"{current_server()})\n"
        "`/help` — 도움말\n"
        "\n"
        "*여러 개 한 번에:* 공백/콤마/줄바꿈 모두 구분자로 가능\n"
        "예) `/add user1 user2,user3` 또는 줄바꿈으로\n"
        "\n"
        "*엑셀 일괄 등록:* .xlsx 파일을 채팅에 첨부하면 A열 2행부터 ID로 추가합니다 "
        "(1행은 헤더, 기존 엑셀 포맷 그대로)"
    )
    await update.message.reply_text(text, parse_mode=ParseMode.MARKDOWN)


@authorized
async def cmd_list(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    ids = id_store.list_ids()
    if not ids:
        await update.message.reply_text("등록된 ID가 없습니다.")
        return
    body = "\n".join(f"{i + 1}. {v}" for i, v in enumerate(ids))
    await update.message.reply_text(f"등록된 ID ({len(ids)}개):\n{body}")


@authorized
async def cmd_add(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    items = _parse_args(update.message.text)
    if not items:
        await update.message.reply_text(
            "사용법: /add <ID> [ID ...]\n공백/콤마/줄바꿈 모두 구분자로 사용 가능."
        )
        return
    added, dupes = id_store.add_ids(items)
    lines = []
    if added:
        lines.append(f"추가됨 ({len(added)}): {', '.join(added)}")
    if dupes:
        lines.append(f"이미 있음 ({len(dupes)}): {', '.join(dupes)}")
    if not lines:
        lines.append("처리된 항목이 없습니다.")
    await update.message.reply_text("\n".join(lines))


@authorized
async def cmd_del(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    items = _parse_args(update.message.text)
    if not items:
        await update.message.reply_text("사용법: /del <ID> [ID ...]")
        return
    removed, missing = id_store.remove_ids(items)
    lines = []
    if removed:
        lines.append(f"삭제됨 ({len(removed)}): {', '.join(removed)}")
    if missing:
        lines.append(f"없음 ({len(missing)}): {', '.join(missing)}")
    if not lines:
        lines.append("처리된 항목이 없습니다.")
    await update.message.reply_text("\n".join(lines))


@authorized
async def cmd_server(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    if not context.args:
        await update.message.reply_text(
            f"현재 서버: {current_server()}\n변경: /server <KR/JP/GLB | CN | ...>"
        )
        return
    new_server = " ".join(context.args).strip()
    state = _load_state()
    state["server"] = new_server
    _save_state(state)
    await update.message.reply_text(f"서버를 '{new_server}'(으)로 변경했습니다.")


@authorized
async def cmd_coupon(update: Update, _: ContextTypes.DEFAULT_TYPE) -> None:
    coupons = _parse_args(update.message.text)
    if not coupons:
        await update.message.reply_text(
            "사용법: /coupon <코드> [코드 ...]\n공백/콤마/줄바꿈 모두 가능."
        )
        return
    ids = id_store.list_ids()
    if not ids:
        await update.message.reply_text(
            "등록된 ID가 없습니다. /add 또는 .xlsx 파일 업로드로 먼저 추가하세요."
        )
        return

    if _run_lock.locked():
        await update.message.reply_text(
            "이미 다른 등록 작업이 진행 중입니다. 잠시 후 다시 시도하세요."
        )
        return

    server = current_server()
    total = len(ids) * len(coupons)
    await update.message.reply_text(
        f"시작합니다. ID {len(ids)}개 × 쿠폰 {len(coupons)}개 = {total}건\n"
        f"서버: {server}\n"
        f"처리가 끝나면 결과를 한 번에 정리해서 보내드립니다."
    )

    def log_progress(line: str) -> None:
        # Console-only progress (no Telegram spam). Useful when watching the
        # bot window during a long run.
        logger.info("[run] %s", line)

    async with _run_lock:
        try:
            results = await asyncio.to_thread(
                coupon_engine.register_coupons,
                ids,
                coupons,
                server,
                HEADLESS,
                log_progress,
            )
        except Exception as e:
            logger.exception("coupon run failed")
            await update.message.reply_text(f"실패: {e}")
            return

    await update.message.reply_text(
        _format_summary(results), parse_mode=ParseMode.MARKDOWN
    )


@authorized
async def handle_excel(update: Update, context: ContextTypes.DEFAULT_TYPE) -> None:
    """Receive an .xlsx file and bulk-add IDs from column A (row 2 onward).

    Matches the format the legacy Tkinter app used so users can keep their
    existing spreadsheets.
    """
    doc = update.message.document
    if doc is None:
        return
    name = (doc.file_name or "").lower()
    if not name.endswith((".xlsx", ".xlsm")):
        await update.message.reply_text("엑셀 파일(.xlsx)만 처리합니다.")
        return

    tmp_path = base_dir() / "_tmp_upload.xlsx"
    try:
        tg_file = await context.bot.get_file(doc.file_id)
        await tg_file.download_to_drive(custom_path=str(tmp_path))
    except Exception as e:
        await update.message.reply_text(f"파일 다운로드 실패: {e}")
        return

    try:
        from openpyxl import load_workbook

        wb = load_workbook(tmp_path, read_only=True, data_only=True)
        ws = wb.active
        ids = [
            str(row[0].value).strip()
            for row in ws.iter_rows(min_row=2)
            if row and row[0].value is not None and str(row[0].value).strip()
        ]
        wb.close()
    except Exception as e:
        logger.exception("excel parse failed")
        await update.message.reply_text(f"엑셀 읽기 실패: {e}")
        return
    finally:
        try:
            tmp_path.unlink(missing_ok=True)
        except Exception:
            pass

    if not ids:
        await update.message.reply_text(
            "엑셀에서 ID를 찾지 못했습니다. A열 2행부터 ID를 채워주세요 (1행은 헤더)."
        )
        return

    added, dupes = id_store.add_ids(ids)
    preview = lambda xs: ", ".join(xs[:20]) + (f" 외 {len(xs) - 20}개" if len(xs) > 20 else "")
    lines = [f"엑셀에서 {len(ids)}건 인식"]
    if added:
        lines.append(f"추가됨 ({len(added)}): {preview(added)}")
    if dupes:
        lines.append(f"이미 있음 ({len(dupes)}): {preview(dupes)}")
    await update.message.reply_text("\n".join(lines))


def _format_summary(results: list[coupon_engine.CouponResult]) -> str:
    total = len(results)
    ok_n = sum(1 for r in results if r.ok)
    fail_n = total - ok_n

    success_by_id: dict[str, Counter] = {}
    failure_by_id: dict[str, Counter] = {}
    for r in results:
        bucket = success_by_id if r.ok else failure_by_id
        bucket.setdefault(r.user_id, Counter())[r.message] += 1

    lines = [f"*완료* — 성공 {ok_n} / 실패 {fail_n} / 총 {total}"]

    if success_by_id:
        lines.append("")
        lines.append("✅ *성공*")
        for uid, counter in success_by_id.items():
            parts = ", ".join(f"{msg}×{n}" for msg, n in counter.items())
            lines.append(f"• `{uid}` — {parts}")

    if failure_by_id:
        lines.append("")
        lines.append("❌ *실패*")
        for uid, counter in failure_by_id.items():
            parts = ", ".join(f"{msg}×{n}" for msg, n in counter.items())
            lines.append(f"• `{uid}` — {parts}")

    return "\n".join(lines)


def main() -> None:
    if not TOKEN:
        raise SystemExit("TELEGRAM_BOT_TOKEN 이 비어있습니다. .env 를 확인하세요.")
    if not ALLOWED_CHAT_IDS:
        raise SystemExit("ALLOWED_CHAT_IDS 가 비어있습니다. .env 를 확인하세요.")

    app = Application.builder().token(TOKEN).build()
    app.add_handler(CommandHandler("start", cmd_start))
    app.add_handler(CommandHandler("help", cmd_help))
    app.add_handler(CommandHandler("list", cmd_list))
    app.add_handler(CommandHandler("add", cmd_add))
    app.add_handler(CommandHandler("del", cmd_del))
    app.add_handler(CommandHandler("server", cmd_server))
    app.add_handler(CommandHandler("coupon", cmd_coupon))
    # Accept .xlsx attachments for bulk ID import.
    app.add_handler(
        MessageHandler(
            filters.Document.FileExtension("xlsx")
            | filters.Document.FileExtension("xlsm"),
            handle_excel,
        )
    )

    logger.info("Bot starting. Allowed chats: %s", ALLOWED_CHAT_IDS)
    app.run_polling(allowed_updates=Update.ALL_TYPES)


if __name__ == "__main__":
    main()
